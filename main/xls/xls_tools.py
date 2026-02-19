import os
import time
from datetime import datetime
from copy import copy

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter

from vetis_api.models import *


STORAGE_DIR = 'main/xls/generated'


def stock_entries_to_xls(enterprise: Enterprise) -> str | None:

    delete_old_files(STORAGE_DIR)

    stock_entries = StockEntry.objects.filter(enterprise=enterprise, is_last=True, is_active=True, volume__gt=0).select_related('main', 'product', 'product_item', 'unit')

    wb = load_workbook(filename=f'main/xls/stock_template.xlsx')

    ws = wb['data']
    table = ws.tables['stock_entries']

    (col_s, row_s, col_e, row_e) = range_boundaries(table.ref)

    first_row_idx = row_s + table.headerRowCount # skip header row(s)

    current_row_idx = row_e + 1

    # s, ХС
    # s, Предприятие
    # s, Номер записи
    # d, Дата поступления
    # s, Статус
    # s, Наименование продукции
    # s, ЕИ
    # n, Объем
    # n, Остаток
    # s, Выработано
    # s, Годен до
    # d, Срок годности
    # s, Источник
    # s, Группа
    # s, Тип продукции

    for stock_entry in stock_entries:
        new_row = [
            str(stock_entry.enterprise.business_entity),
            str(stock_entry.enterprise),
            str(stock_entry.entry_number),
            stock_entry.main.date_created.date() if stock_entry.main.date_created else None,
            stock_entry.main.get_initial_status_display(),
            stock_entry.product_item_name,
            str(stock_entry.unit),
            stock_entry.main.initial_volume,
            stock_entry.volume,
            stock_entry.date_produced_display,
            stock_entry.date_expiry_display,
            stock_entry.date_expiry.astimezone(tz=TZ_MOSCOW).date(),
            stock_entry.main.source_ent_name,
            stock_entry.product_item.assort_group.name if stock_entry.product_item.assort_group else '! НЕ УКАЗАНА',
            stock_entry.product.name
        ]

        ws.append(new_row)

        for target_cell, source_cell in zip(ws[current_row_idx], ws[first_row_idx]):
            target_cell.number_format = copy(source_cell.number_format)
        
        current_row_idx += 1

    last_row_idx = current_row_idx - 1

    ws.delete_rows(first_row_idx) # deleting first sample row
    last_row_idx -= 1

    table.ref = f'{get_column_letter(col_s)}{row_s}:{get_column_letter(col_e)}{last_row_idx}'

    ws['B1'] = enterprise.stock_entries_last_updated.astimezone(tz=TZ_MOSCOW).replace(tzinfo=None)

    pivot = wb['Сводная']._pivots[0]
    pivot.cache.refreshOnLoad = True

    filename = f'stock_entries_{datetime.now(tz=TZ_MOSCOW).strftime('%y%m%d_%H%M%S')}.xlsx'

    fullpath = f'{STORAGE_DIR}/{filename}'

    wb.save(fullpath)

    return fullpath


def delete_old_files(folder_path: str, hours: int = 1) -> int:
    if not os.path.exists(folder_path):
        return 0
    
    cutoff_time = time.time() - (hours * 60 * 60)

    to_delete = []

    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        
        if os.path.isfile(file_path):
            
            file_mtime = os.path.getmtime(file_path)
            
            if file_mtime < cutoff_time:
                to_delete.append((filename, file_path, file_mtime))
    
    deleted_count = 0
    for filename, file_path, _ in to_delete:
        try:
            os.remove(file_path)
            deleted_count += 1
        except Exception as e:
            print(f"Error deleting {filename}: {e}")
    
    return deleted_count
